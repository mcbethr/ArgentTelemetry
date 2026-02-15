#!/usr/bin/env python3
"""
Create one Excel workbook containing all telemetry rows from authors who commented
on 2+ different channel_name values.

Input:
  Master .xlsx with sheet: "Telemetry Data"

Output:
  CrossChannelAuthors-YYYY-MM-DD_HHMMSS.xlsx (in outdir)

Usage:
  python3 export_cross_channel_authors.py \
    --infile "Master-2026-02-14_224319.xlsx" \
    --outdir "outputs" \
    --min-channels 2
"""

import os
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo
from collections import defaultdict
from typing import Any, Dict, List, Set, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def autosize_columns(ws, max_width: int = 80):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def build_header_index(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, name in enumerate(header_row):
        if name is None:
            continue
        idx[str(name).strip()] = i
    return idx


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--infile", required=True, help="Path to Master .xlsx")
    ap.add_argument("--sheet", default="Telemetry Data", help="Sheet name (default: Telemetry Data)")
    ap.add_argument("--outdir", default="outputs", help="Output directory")
    ap.add_argument("--author-col", default="author", help="Column name for author (default: author)")
    ap.add_argument("--channel-col", default="channel_name", help="Column name for channel (default: channel_name)")
    ap.add_argument("--min-channels", type=int, default=2, help="Minimum distinct channels (default: 2)")
    args = ap.parse_args()

    os.makedirs(args.outdir, exist_ok=True)

    # Load master workbook
    wb = load_workbook(args.infile, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")

    ws = wb[args.sheet]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        raise SystemExit("Telemetry Data sheet is empty.")

    col_index = build_header_index(header)

    if args.author_col not in col_index:
        raise SystemExit(f"Author column '{args.author_col}' not found in header.")
    if args.channel_col not in col_index:
        raise SystemExit(f"Channel column '{args.channel_col}' not found in header.")

    author_i = col_index[args.author_col]
    channel_i = col_index[args.channel_col]

    # First pass: gather channel sets per author + store all rows
    channels_by_author: Dict[str, Set[str]] = defaultdict(set)
    rows_by_author: Dict[str, List[Tuple[Any, ...]]] = defaultdict(list)

    total_rows = 0
    for r in rows_iter:
        total_rows += 1

        author = r[author_i]
        if author is None or str(author).strip() == "":
            continue
        author_s = str(author).strip()

        channel = r[channel_i]
        channel_s = str(channel).strip() if channel is not None else ""

        rows_by_author[author_s].append(r)
        if channel_s:
            channels_by_author[author_s].add(channel_s)

    wb.close()

    # Determine cross-channel authors
    cross_authors = {
        a for a, chans in channels_by_author.items()
        if len(chans) >= args.min_channels
    }

    print(f"Loaded {total_rows} rows.")
    print(f"Found {len(cross_authors)} authors with >= {args.min_channels} unique channel_name values.")

    # Build output workbook
    tz = ZoneInfo("America/New_York")
    stamp = datetime.now(tz).strftime("%Y-%m-%d_%H%M%S")
    out_path = os.path.join(args.outdir, f"CrossChannelAuthors-{stamp}.xlsx")

    out_wb = Workbook()

    # Sheet 1: concatenated telemetry for cross-channel authors
    out_ws = out_wb.active
    out_ws.title = "Cross-Channel Telemetry"
    out_ws.append(list(header))

    written = 0
    for author in sorted(cross_authors, key=str.lower):
        for r in rows_by_author.get(author, []):
            out_ws.append(list(r))
            written += 1

    # Sheet 2: index
    idx_ws = out_wb.create_sheet("Author Index")
    idx_ws.append(["author", "unique_channel_count", "channels"])

    for author in sorted(cross_authors, key=str.lower):
        chans = sorted(list(channels_by_author.get(author, set())))
        idx_ws.append([author, len(chans), ", ".join(chans)])

    autosize_columns(out_ws)
    autosize_columns(idx_ws)

    out_wb.save(out_path)

    print(f"Wrote {written} telemetry rows to: {out_path}")


if __name__ == "__main__":
    main()