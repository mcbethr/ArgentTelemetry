#!/usr/bin/env python3
"""
Export YouTube comments to CSV:
- author (display name)
- publishedAt
- likeCount
- isPinned (top-level comments only)

Requires:
  pip install google-api-python-client python-dateutil

Usage:
  export YT_API_KEY="Key Here"
  python Argent.py "https://www.youtube.com/watch?v=2erRc2mimhE" --out comments.csv --include-replies
  Batch mode: python3 Argent.py --video-file videos.txt --out-dir exports --include-replies
  """

import os
import re
import csv
import argparse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any, Iterator, Optional, List
from dotenv import load_dotenv, find_dotenv
load_dotenv()
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

def read_video_list(path: str) -> List[str]:
    """
    Reads newline-separated video URLs/IDs.

    Rules:
      - Blank lines ignored
      - Lines whose first non-whitespace char is '#' are ignored
      - Inline comments supported: anything after a ' #' is ignored
        Example: https://youtu.be/abc123xyz00  # note
    """
    vids: List[str] = []
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue

            # Full-line comment (allow leading whitespace)
            if line.startswith("#"):
                continue

            # Inline comment: split on first '#' and keep left side
            # This lets you annotate URLs on the same line.
            if "#" in line:
                line = line.split("#", 1)[0].strip()

            if not line:
                continue

            vids.append(line)

    return vids


def export_video_to_excel(yt, video_input: str, out_dir: str, order: str, include_replies: bool) -> str:
    """
    Exports one video's comments to an .xlsx and returns the output path.
    """
    video_id = extract_video_id(video_input)

    meta = get_video_metadata(yt, video_id)
    channel_handle = get_channel_handle_like(yt, meta["channel_id"])

    channel_name = meta["channel_name"]
    channel_id = meta["channel_id"]
    safe_channel = sanitize_filename(channel_name)

    # Make filename unique per video
    out_xlsx = os.path.join(out_dir, f"{safe_channel}-{channel_id}-{video_id}.xlsx")

    wb = Workbook()

    # Sheet 1: Telemetry Data
    ws_telemetry = wb.active
    ws_telemetry.title = "Telemetry Data"
    ws_telemetry.append([
        "video_id",
        "comment_id",
        "parent_id",
        "author",
        "author_channel_url",
        "published_at",
        "updated_at",
        "like_count",
        "is_pinned",
        "text",
    ])

    count = 0
    for row in iter_comment_threads(yt, video_id, order=order, include_replies=include_replies):
        ws_telemetry.append([
            video_id,
            row.get("comment_id", ""),
            row.get("parent_id", ""),
            row.get("author", ""),
            row.get("author_channel_url", ""),
            row.get("published_at", ""),
            row.get("updated_at", ""),
            row.get("like_count", 0),
            row.get("is_pinned", False),
            row.get("text", ""),
        ])
        count += 1

    # Sheet 2: Comment data
    ws_meta = wb.create_sheet("Comment data")
    ws_meta.append(["Channel name", "Channel handle", "Video link", "Video Title", "Upload date"])
    ws_meta.append([
        meta["channel_name"],
        channel_handle,
        meta["video_link"],
        meta["video_title"],
        meta["upload_date"],
    ])

    autosize_columns(ws_telemetry)
    autosize_columns(ws_meta)

    wb.save(out_xlsx)
    print(f"Wrote {count} comments -> {out_xlsx}")
    return out_xlsx

def get_video_metadata(yt, video_id: str) -> Dict[str, str]:
    """
    Returns video metadata needed for the 'Channel Data' tab.
    """
    resp = yt.videos().list(
        part="snippet",
        id=video_id
    ).execute()

    if not resp.get("items"):
        raise ValueError(f"Video not found or not accessible: {video_id}")

    snip = resp["items"][0]["snippet"]
    return {
        "video_id": video_id,
        "video_title": snip.get("title", ""),
        "upload_date": snip.get("publishedAt", ""),
        "channel_id": snip.get("channelId", ""),
        "channel_name": snip.get("channelTitle", ""),
        "video_link": f"https://www.youtube.com/watch?v={video_id}",
    }


def get_channel_handle_like(yt, channel_id: str) -> str:
    """
    Best-effort channel 'handle' field. The API commonly provides 'customUrl'
    which is often '@handle' but not guaranteed.
    """
    resp = yt.channels().list(
        part="snippet",
        id=channel_id
    ).execute()

    if not resp.get("items"):
        return ""

    snip = resp["items"][0]["snippet"]
    # Often looks like "@RyanMcBeth" or "c/SomeName" depending on channel
    return snip.get("customUrl", "") or ""


def autosize_columns(ws):
    """
    Simple autosize for openpyxl based on max string length per column.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

def get_channel_info(yt, video_id: str):
    """
    Given a video ID, returns (channel_name, channel_id)
    """
    req = yt.videos().list(
        part="snippet",
        id=video_id
    )
    resp = req.execute()
    snippet = resp["items"][0]["snippet"]
    return snippet["channelTitle"], snippet["channelId"]


def sanitize_filename(name: str) -> str:
    """
    Makes channel name safe for macOS filenames
    """
    return re.sub(r'[<>:"/\\|?*]', '', name).strip()

def extract_video_id(url_or_id: str) -> str:
    # Accept raw video ID
    if re.fullmatch(r"[A-Za-z0-9_-]{11}", url_or_id):
        return url_or_id

    # Common URL formats
    patterns = [
        r"v=([A-Za-z0-9_-]{11})",
        r"youtu\.be/([A-Za-z0-9_-]{11})",
        r"shorts/([A-Za-z0-9_-]{11})",
        r"embed/([A-Za-z0-9_-]{11})",
    ]
    for p in patterns:
        m = re.search(p, url_or_id)
        if m:
            return m.group(1)

    raise ValueError(f"Could not extract a video id from: {url_or_id}")


def youtube_client(api_key: str):
    return build("youtube", "v3", developerKey=api_key, cache_discovery=False)


def iter_comment_threads(
    yt,
    video_id: str,
    order: str = "time",
    include_replies: bool = False,
) -> Iterator[Dict[str, Any]]:
    """
    Yields dict rows for each top-level comment (and optionally replies).
    """
    page_token: Optional[str] = None

    while True:
        req = yt.commentThreads().list(
            part="snippet,replies",
            videoId=video_id,
            maxResults=100,
            order=order,          # "time" or "relevance"
            textFormat="plainText",
            pageToken=page_token,
        )
        resp = req.execute()

        for item in resp.get("items", []):
            thr_snip = item["snippet"]
            top = thr_snip["topLevelComment"]["snippet"]

            # Pinned is available on the commentThread snippet in many cases
            is_pinned = thr_snip.get("isPinned")

            yield {
                "comment_id": item["snippet"]["topLevelComment"]["id"],
                "parent_id": "",
                "author": top.get("authorDisplayName"),
                "author_channel_url": top.get("authorChannelUrl"),
                "published_at": top.get("publishedAt"),
                "updated_at": top.get("updatedAt"),
                "like_count": top.get("likeCount", 0),
                "is_pinned": is_pinned,
                "text": top.get("textDisplay", ""),
            }

            if include_replies and "replies" in item:
                replies = item["replies"].get("comments", [])
                for r in replies:
                    r_snip = r["snippet"]
                    yield {
                        "comment_id": r["id"],
                        "parent_id": item["snippet"]["topLevelComment"]["id"],
                        "author": r_snip.get("authorDisplayName"),
                        "author_channel_url": r_snip.get("authorChannelUrl"),
                        "published_at": r_snip.get("publishedAt"),
                        "updated_at": r_snip.get("updatedAt"),
                        "like_count": r_snip.get("likeCount", 0),
                        "is_pinned": False,
                        "text": r_snip.get("textDisplay", ""),
                    }

        page_token = resp.get("nextPageToken")
        if not page_token:
            break

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("video", nargs="?", help="YouTube video URL or 11-char video id (single mode)")
    ap.add_argument("--video-file", help="Text file with YouTube video URLs/IDs, one per line (batch mode)")
    ap.add_argument("--out-dir", default="exports", help="Output directory for batch exports")
    ap.add_argument("--order", choices=["time", "relevance"], default="time", help="API sort order")
    ap.add_argument("--include-replies", action="store_true", help="Also export replies")
    args = ap.parse_args()

    load_dotenv()
    api_key = os.getenv('YT_API_KEY')
    if not api_key:
        raise SystemExit("Set env var YT_API_KEY to your YouTube Data API key first.")

    yt = youtube_client(api_key)

    # ======================
    # Batch mode
    # ======================
    if args.video_file:
        os.makedirs(args.out_dir, exist_ok=True)
        videos = read_video_list(args.video_file)

        if not videos:
            raise SystemExit(f"No videos found in {args.video_file}")

        ok, failed = 0, 0
        for v in videos:
            try:
                export_video_to_excel(
                    yt=yt,
                    video_input=v,
                    out_dir=args.out_dir,
                    order=args.order,
                    include_replies=args.include_replies,
                )
                ok += 1
            except Exception as e:
                print(f"[!] Failed for {v}: {e}")
                failed += 1

        print(f"\nDone. Success: {ok}, Failed: {failed}")
        return

    # ======================
    # Single video mode
    # ======================
    if not args.video:
        raise SystemExit("Provide a single video URL/ID or use --video-file <path>.")

    video_id = extract_video_id(args.video)

    export_video_to_excel(
        yt=yt,
        video_input=video_id,
        out_dir=".",
        order=args.order,
        include_replies=args.include_replies,
    )
    ap = argparse.ArgumentParser()
    ap.add_argument("video", nargs="?", help="YouTube video URL or 11-char video id (single mode)")
    ap.add_argument("--video-file", help="Text file with YouTube video URLs/IDs, one per line (batch mode)")
    ap.add_argument("--out-dir", default="exports", help="Output directory for batch exports")
    ap.add_argument("--order", choices=["time", "relevance"], default="time", help="API sort order")
    ap.add_argument("--include-replies", action="store_true", help="Also export replies")
    args = ap.parse_args()

    load_dotenv() #load the environment variables.

    api_key = os.getenv('YT_API_KEY')
    if not api_key:
        raise SystemExit("Set env var YT_API_KEY to your YouTube Data API key first.")

    video_id = extract_video_id(args.video)
    yt = youtube_client(api_key)

 # Batch mode
    if not args.video:
      raise SystemExit("Provide a single video URL/ID or use --video-file <path>.")

    if args.video_file:
        os.makedirs(args.out_dir, exist_ok=True)
        videos = read_video_list(args.video_file)

        if not videos:
            raise SystemExit(f"No videos found in {args.video_file}")

        ok, failed = 0, 0
        for v in videos:
            try:
                export_video_to_excel(
                    yt=yt,
                    video_input=v,
                    out_dir=args.out_dir,
                    order=args.order,
                    include_replies=args.include_replies,
                )
                ok += 1
            except HttpError as e:
                print(f"[!] API error for {v}: {e}")
                failed += 1
            except Exception as e:
                print(f"[!] Failed for {v}: {e}")
                failed += 1

        print(f"Done. Success: {ok}, Failed: {failed}. Output folder: {args.out_dir}")
        return

    # Metadata for the "Channel Data" sheet
    meta = get_video_metadata(yt, video_id)
    channel_handle = get_channel_handle_like(yt, meta["channel_id"])

    channel_name, channel_id = get_channel_info(yt, video_id)
    safe_channel = sanitize_filename(channel_name)

    output_filename = f"{safe_channel}-{channel_id}.csv"

    print(f"Output file: {output_filename}")

    fieldnames = [
        "comment_id",
        "parent_id",
        "author",
        "author_channel_url",
        "published_at",
        "updated_at",
        "like_count",
        "is_pinned",
        "text",
    ]

    try:
        wb = Workbook()

        # Sheet 1: Telemetry Data (comments)
        ws_telemetry = wb.active
        ws_telemetry.title = "Telemetry Data"

        telemetry_headers = [
            "video_id",
            "comment_id",
            "parent_id",
            "author",
            "author_channel_url",
            "published_at",
            "updated_at",
            "like_count",
            "is_pinned",
            "text",
        ]
        ws_telemetry.append(telemetry_headers)

        count = 0
        for row in iter_comment_threads(yt, video_id, order=args.order, include_replies=args.include_replies):
            ws_telemetry.append([
                video_id,
                row.get("comment_id", ""),
                row.get("parent_id", ""),
                row.get("author", ""),
                row.get("author_channel_url", ""),
                row.get("published_at", ""),
                row.get("updated_at", ""),
                row.get("like_count", 0),
                row.get("is_pinned", False),
                row.get("text", ""),
            ])
            count += 1

        # Sheet 2: Channel Data (channel + video metadata)
        ws_meta = wb.create_sheet("Channel Data")
        meta_headers = ["Channel name", "Channel handle", "Video link", "Video Title", "Upload date"]
        ws_meta.append(meta_headers)
        ws_meta.append([
            meta["channel_name"],
            channel_handle,
            meta["video_link"],
            meta["video_title"],
            meta["upload_date"],
        ])

        autosize_columns(ws_telemetry)
        autosize_columns(ws_meta)

        # Output filename: <ChannelName>-<ChannelID>.xlsx
        out_xlsx = f"{safe_channel}-{channel_id}.xlsx"

        wb.save(out_xlsx)
        print(f"Wrote {count} comments to {out_xlsx}")

    except HttpError as e:
        raise SystemExit(f"YouTube API error: {e}")

if __name__ == "__main__":
    main()
